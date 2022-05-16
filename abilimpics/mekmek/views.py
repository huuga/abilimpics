from django.shortcuts import render
import openpyxl




def index(request):
    result_dict = dict()
    if request.method == 'POST':
        wb = openpyxl.load_workbook('Отчёт трудоустройства - Республика Саха (Якутия) - За 2018-2019-2020-2021.xlsx')
        ws = wb.active


        search_string = request.POST.get('full_name')
        partic_info = list()
        for i in range(6, ws.max_row + 1):
            if ws.cell(i, 3).value:
                if search_string.lower().strip() in ws.cell(i, 3).value.lower().strip():
                    partic_info.append('******************************************')
                    partic_info.append(f'{ws.cell(i, 3).value.upper()} ({ws.cell(i, 4).value})')
                    partic_info.append(f'Заболевание - {ws.cell(i, 20).value}')
                    partic_info.append(f"Обучается в данное время - {ws.cell(i, 13).value}")
                    if ws.cell(i, 13).value.strip().lower() == 'да':
                        partic_info.append(f'----Уровень образования - {ws.cell(i, 14).value}')
                        partic_info.append(f'----Планируемый год окончания образовательной организации - {ws.cell(i, 15).value}')
                        partic_info.append(f'----Название образовательной организации - {ws.cell(i, 16).value}')
                        partic_info.append(f'----Специальность - {ws.cell(i, 17).value}')
                        partic_info.append(f'----Гарантированное трудоустройство после обучения - {ws.cell(i, 39).value}')
                        partic_info.append(f'----Целевое обучение с последующим трудоустройством - {ws.cell(i, 41).value}')
                    partic_info.append(f"Трудоустроен - {ws.cell(i, 23).value}")
                    if ws.cell(i, 23).value.strip().lower() == 'нет':
                        partic_info.append(f'----Причина нетрудоустройства - {ws.cell(i, 32).value}')
                        partic_info.append(f'----На учете в органах службы занятости - {ws.cell(i, 33).value}')
                        partic_info.append(f'----Повышение квалификации пройдено - {ws.cell(i, 36).value}')
                        partic_info.append(f'----Размещено резюме на портале "Работа в России" - {ws.cell(i, 38).value}')
                    else:
                        partic_info.append(f'----Наименование организации - {ws.cell(i, 24).value}')
                        partic_info.append(f'----Должность - {ws.cell(i, 25).value}')
                        partic_info.append(f'----Дата приема на работу - {ws.cell(i, 26).value}')
                    partic_info.append(f'Проходит стажировку - {ws.cell(i, 28).value}')
                    result_dict['shalom'] = partic_info
    return render(request, 'mekmek/index.html', context=result_dict)



